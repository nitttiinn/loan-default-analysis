import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import os

#1. Load Analyzed Data:
df = pd.read_csv('Data/analyzed_loan_data.csv')

#2. Pre-Calculate all analysis:
# Default Rate Overall
total_loans = len(df)
total_defaults = df['Status'].sum()
default_rate = (total_defaults/total_loans*100).round(2)
paid_back = total_loans-total_defaults

# By Credit Score:
credit_default = (df.groupby('Credit_Score_Group')['Status'].mean()*100).round(2)

# By Income Group:
income_order = ['Very Low','Low','Medium','High','Very High']
income_default = (df.groupby('income_group')['Status'].mean()*100).round(2)
income_default = income_default.reindex(income_order) 

# By Loan Amount:
loan_order = ['Very Small','Small','Medium','Large','Very Large']
loan_amount_default = (df.groupby('loan_amount_group')['Status'].mean()*100).round(2)
loan_amount_default = loan_amount_default.reindex(loan_order)

# By LTV:
ltv_order = ['Very Low','Low','Medium','High','Very High']
ltv_default = (df.groupby('ltv_group')['Status'].mean()*100).round(2)
ltv_default = ltv_default.reindex(ltv_order)

# By Region:
region_default = (df.groupby('Region')['Status'].mean()*100).round(2).sort_values(ascending=False)

# By Loan Purpose:
loan_purpose_default = (df.groupby('loan_purpose')['Status'].mean()*100).round(2)

# Summary Stats:
summary = df.groupby('Status')[['loan_amount','income','Credit_Score','LTV']].mean().round(2)
summary.index = ['Paid Back','Defaulted']

#3. Helper Function:
def style_header(ws,row_num,num_cols,color = '1F4E79'):
    for col in range(1,num_cols+1):
        cell = ws.cell(row=row_num,column = col)
        cell.font = Font(bold=True,color='FFFFFF',size = 11)
        cell.fill = PatternFill(fill_type='solid',fgColor=color)
        cell.alignment = Alignment(horizontal='center',vertical='center')

def style_title(ws, cell_ref, text, size=16):
    ws[cell_ref] = text
    ws[cell_ref].font = Font(bold=True, size=size, color="1F4E79")

def write_dataframe(ws, df, start_row, header_color="1F4E79"):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    style_header(ws, start_row, len(df.columns) + 1, header_color)


#4. Create Workbook:
wb = Workbook()

# Sheet 1 -> Executive Summary
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 20

# Title:
style_title(ws1, "A1", "🏦 Loan Default Analysis Report", size=18)
ws1["A2"] = "Executive Summary"
ws1["A2"].font = Font(size=12, color="666666")

# KPI Section:
ws1['A4'] = "KEY METRICS"
ws1['A4'].font = Font(bold=True,size = 13,color="1F4E79")

kpis = [
    ("Total Loans Analyzed", f"{total_loans:,}"),
    ("Total Defaults", f"{total_defaults:,}"),
    ("Total Paid Back", f"{paid_back:,}"),
    ("Overall Default Rate", f"{default_rate}%"),
    ("Strongest Risk Factor", "LTV Ratio"),
    ("Riskiest Region", "North-East (30.4%)"),
    ("Safest Income Group", "Medium (18.92%)"),
    ("Riskiest Loan Amount", "Very Large (48.88%)"),
]

for i, (metric, value) in enumerate(kpis, start=5):
    ws1.cell(row=i, column=1, value=metric).font = Font(bold=True)
    ws1.cell(row=i, column=2, value=value).font = Font(color="1F4E79", bold=True)

# recommendations
ws1["A14"] = "KEY RECOMMENDATIONS"
ws1["A14"].font = Font(bold=True, size=13, color="1F4E79")

recommendations = [
    "1. Strictly limit loans with Very High LTV — 65% default rate!",
    "2. Apply stricter approval for Very Large loans — 48.88% default rate",
    "3. Investigate Loan Purpose P2 — highest default rate at 32.98%",
    "4. Use Income + LTV together — Credit Score alone is weak predictor",
    "5. North-East region needs stricter loan approval policies",
    "6. Joint applications are safest — encourage co-applicants",
]

for i, rec in enumerate(recommendations, start=15):
    ws1.cell(row=i, column=1, value=rec).font = Font(size=11)

# =====================
# Sheet 2 — Default by Credit Score
# =====================
ws2 = wb.create_sheet("Credit Score Analysis")
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 20

style_title(ws2, "A1", "Default Rate by Credit Score Group")
write_dataframe(ws2, credit_default.to_frame(), start_row=3)

# add chart image
if os.path.exists('charts/chart1_credit_score.png'):
    img = Image('charts/chart1_credit_score.png')
    img.width = 500
    img.height = 250
    ws2.add_image(img, 'D3')

# =====================
# Sheet 3 — Income Analysis
# =====================
ws3 = wb.create_sheet("Income Analysis")
ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 20

style_title(ws3, "A1", "Default Rate by Income Group")
write_dataframe(ws3, income_default.to_frame(), start_row=3)

if os.path.exists('charts/chart3_income_group.png'):
    img = Image('charts/chart3_income_group.png')
    img.width = 500
    img.height = 250
    ws3.add_image(img, 'D3')

# =====================
# Sheet 4 — LTV Analysis
# =====================
ws4 = wb.create_sheet("LTV Analysis")
ws4.column_dimensions['A'].width = 25
ws4.column_dimensions['B'].width = 20

style_title(ws4, "A1", "Default Rate by LTV Group")
write_dataframe(ws4, ltv_default.to_frame(), start_row=3)

if os.path.exists('charts/chart5_ltv.png'):
    img = Image('charts/chart5_ltv.png')
    img.width = 500
    img.height = 250
    ws4.add_image(img, 'D3')

# =====================
# Sheet 5 — Regional Analysis
# =====================
ws5 = wb.create_sheet("Regional Analysis")
ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 20

style_title(ws5, "A1", "Default Rate by Region")
write_dataframe(ws5, region_default.to_frame(), start_row=3)

if os.path.exists('charts/chart6_region.png'):
    img = Image('charts/chart6_region.png')
    img.width = 500
    img.height = 250
    ws5.add_image(img, 'D3')

# =====================
# Sheet 6 — Loan Purpose Analysis
# =====================
ws6 = wb.create_sheet("Loan Purpose Analysis")
ws6.column_dimensions['A'].width = 25
ws6.column_dimensions['B'].width = 20

style_title(ws6, "A1", "Default Rate by Loan Purpose")
write_dataframe(ws6, loan_purpose_default.to_frame(), start_row=3)

if os.path.exists('charts/chart2_loan_purpose.png'):
    img = Image('charts/chart2_loan_purpose.png')
    img.width = 500
    img.height = 250
    ws6.add_image(img, 'D3')

# =====================
# Sheet 7 — Summary Stats
# =====================
ws7 = wb.create_sheet("Summary Statistics")
ws7.column_dimensions['A'].width = 25
ws7.column_dimensions['B'].width = 20
ws7.column_dimensions['C'].width = 20
ws7.column_dimensions['D'].width = 20
ws7.column_dimensions['E'].width = 20

style_title(ws7, "A1", "Average Stats — Defaulted vs Paid Back")
write_dataframe(ws7, summary, start_row=3)

# =====================
# Save report
# =====================
wb.save('Loan_Default_Report.xlsx')
print("Report saved as Loan_Default_Report.xlsx! ✅")

